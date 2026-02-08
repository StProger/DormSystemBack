from __future__ import annotations
from io import BytesIO
from typing import Any
from openpyxl import load_workbook

class ExcelError(Exception): ...

def read_sheet_rows_xlsx(data: bytes, sheet_name: str) -> list[dict[str, Any]]:
    """
    Reads first row as headers. Returns list of dicts for subsequent rows.
    Empty trailing rows are ignored.
    """
    try:
        wb = load_workbook(filename=BytesIO(data), data_only=True)
    except Exception as e:
        raise ExcelError(f"Invalid xlsx: {e}")

    if sheet_name not in wb.sheetnames:
        raise ExcelError(f"Sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not any(headers):
        raise ExcelError("Header row is empty")

    out: list[dict[str, Any]] = []
    for i, r in enumerate(rows[1:], start=2):  # Excel row index
        if r is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        obj = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            obj[key] = r[idx] if idx < len(r) else None
        out.append({"__row__": i, **obj})
    return out
