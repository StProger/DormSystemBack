from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.utils.ru_labels import ru, GUEST_STATUS_RU, TICKET_TYPE_RU, TICKET_STATUS_RU


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 55)

def wb_to_bytes(wb: Workbook) -> bytes:
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def occupancy_to_wb(summary: dict, rooms: list[dict]) -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Метрика", "Значение"])
    ws.append(["Комнат всего", summary["rooms_total"]])
    ws.append(["Вместимость всего", summary["capacity_total"]])
    ws.append(["Заселено", summary["occupied_total"]])
    ws.append(["Свободно", summary["free_total"]])
    ws.append(["Загрузка (%)", round(summary["occupancy_rate"] * 100, 2)])
    _autosize(ws)

    ws2 = wb.create_sheet("Rooms")
    ws2.append(["Номер", "Вместимость", "Заселено", "Свободно", "Последняя оценка", "Комментарий", "Дата оценки"])
    for r in rooms:
        ws2.append([
            r["number"], r["capacity"], r["occupied"], r["free"],
            r.get("last_rating"), r.get("last_rating_comment"), r.get("last_rating_at"),
        ])
    _autosize(ws2)
    return wb

def tickets_to_wb(meta: dict, by_status: list[dict], by_type: list[dict], by_day: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Meta"
    ws.append(["Параметр", "Значение"])
    ws.append(["from", meta.get("from_date")])
    ws.append(["to", meta.get("to_date")])
    _autosize(ws)

    ws1 = wb.create_sheet("ByStatus")
    ws1.append(["Статус", "Кол-во"])
    for x in by_status:
        ws1.append([ru(TICKET_STATUS_RU, x["key"]), x["count"]])
    _autosize(ws1)

    ws2 = wb.create_sheet("ByType")
    ws2.append(["Тип", "Кол-во"])
    for x in by_type:
        ws2.append([ru(TICKET_TYPE_RU, x["key"]), x["count"]])
    _autosize(ws2)

    ws3 = wb.create_sheet("ByDay")
    ws3.append(["День", "Кол-во"])
    for x in by_day:
        ws3.append([x["day"], x["count"]])
    _autosize(ws3)

    return wb

def guest_pass_to_wb(meta: dict, by_status: list[dict], by_day: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Meta"
    ws.append(["Параметр", "Значение"])
    ws.append(["from", meta.get("from_date")])
    ws.append(["to", meta.get("to_date")])
    _autosize(ws)

    ws1 = wb.create_sheet("ByStatus")
    ws1.append(["Статус", "Кол-во"])
    for x in by_status:
        ws1.append([ru(GUEST_STATUS_RU, x["key"]), x["count"]])
    _autosize(ws1)

    ws2 = wb.create_sheet("ByDay")
    ws2.append(["День", "Кол-во"])
    for x in by_day:
        ws2.append([x["day"], x["count"]])
    _autosize(ws2)
    return wb
